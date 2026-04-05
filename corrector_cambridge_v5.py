"""
Cambridge Writing Corrector — motor desacoplado (v5)
=====================================================
Este archivo es el MOTOR. No tiene main(), no tiene input(), no escribe en disco.
Toda la E/S ocurre en memoria (BytesIO). Es importado por app.py.

Estructura del proyecto esperada:
  /tu_proyecto/
    corrector_cambridge_v5.py   ← este archivo
    app.py                       ← interfaz Streamlit
    redacciones/
      Rubrica/
        rubrica.pdf
      Essay/
        Essay.pdf
      Report/
        Report.pdf
      Proposal/
        Proposal.pdf
      Letter/
        Letter.pdf
"""

import io
import os
import re
import time
from pathlib import Path

import pdfplumber
from openai import OpenAI
from google import genai

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────

MODEL_NAME          = "llama-3.3-70b-versatile"
DELAY_BETWEEN_CALLS = 5
LLAMA_BASE_URL      = "https://api.groq.com/openai/v1"

# Ruta base de los PDFs de referencia — relativa a este archivo
_HERE         = Path(__file__).parent
REDACCIONES   = _HERE / "redacciones"
RUBRIC_PATH   = REDACCIONES / "Rubrica" / "rubrica.pdf"

# Columnas del Excel (índice 1-based de openpyxl)
COL_NOMBRE    = 4   # D
COL_REDACCION = 5   # E
COL_GRADE     = 6   # F — Grade Summary
COL_NOTA      = 7   # G — Nota /20
COL_ERRORS    = 8   # H — Error Analysis
COL_COMMENTS  = 9   # I — Examiner Comments
COL_CORRECTED = 10  # J — Corrected Essay
COL_FULL      = 11  # K — Corrección completa


# ─────────────────────────────────────────────────────────────
# CLIENTE API  (se puede sobreescribir desde app.py)
# ─────────────────────────────────────────────────────────────


def get_essay_types() -> list[str]:
    """
    Devuelve los tipos de redacción disponibles:
    carpetas dentro de redacciones/ que tienen su PDF de ejemplo.
    """
    if not REDACCIONES.exists():
        return []
    return sorted(
        d.name for d in REDACCIONES.iterdir()
        if d.is_dir()
        and d.name.lower() != "rubrica"
        and (d / f"{d.name}.pdf").exists()
    )


# ─────────────────────────────────────────────────────────────
# UTILIDADES PDF
# ─────────────────────────────────────────────────────────────

def extract_pdf_text(path: Path) -> str:
    """Extrae texto de un PDF página a página."""
    parts = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
    return "\n\n".join(parts)


# ─────────────────────────────────────────────────────────────
# PROMPT DEL SISTEMA
# ─────────────────────────────────────────────────────────────

SYSTEM_INSTRUCTION = (
    "You are an expert Cambridge Examiner for C1 Advanced and C2 Proficiency levels. "
    "Your ONLY references for grading are the official rubric and correction example provided. "
    "Apply the rubric criteria consistently and literally across every essay you correct.\n\n"

    "For each essay you receive, produce the following sections exactly:\n\n"

    "## GRADE SUMMARY\n"
    "Score each criterion on the 0-5 scale defined in the rubric, "
    "followed by a one-sentence justification. Format:\n"
    "  Content: X/5 — [justification]\n"
    "  Communicative Achievement: X/5 — [justification]\n"
    "  Organization: X/5 — [justification]\n"
    "  Language: X/5 — [justification]\n"
    "  TOTAL: X/20\n\n"

    "## ERROR ANALYSIS\n"
    "List every significant error found. For each error:\n"
    "  * Quote the original fragment in [brackets]\n"
    "  * State the error type (Grammar / Vocabulary / Spelling / Punctuation / "
    "Register / Coherence / Other)\n"
    "  * Provide the corrected form\n"
    "  * Give a brief explanation (one sentence)\n\n"

    "## EXAMINER COMMENTS\n"
    "2 to 4 sentences of global feedback: strengths, main weaknesses, "
    "and the most important area to improve.\n\n"

    "## CORRECTED ESSAY\n"
    "A full rewrite of the essay correcting ALL mistakes while keeping "
    "the student's original ideas, structure, and approximate length. "
    "This section must be self-contained — do not reference the original text.\n\n"

    "CONSISTENCY RULES (critical):\n"
    "- A score of 3/5 in Language means the same thing in every essay.\n"
    "- Do not reward or penalise based on essay length alone.\n"
    "- If two essays have equivalent quality, they must receive identical scores.\n"
    "- Errors must always be classified in the same way regardless of the student."
)


# ─────────────────────────────────────────────────────────────
# CLASE PRINCIPAL
# ─────────────────────────────────────────────────────────────

class CambridgeCorrector:
    def __init__(self, essay_type: str, provider: str, api_key: str):
        self.essay_type = essay_type
        self.provider = provider  # "groq" o "gemini"
        self.api_key = api_key

        # Lógica de carga de PDFs (reintegrada)
        example_path = REDACCIONES / essay_type / f"{essay_type}.pdf"
        if not RUBRIC_PATH.exists():
            raise FileNotFoundError(f"Rúbrica no encontrada: {RUBRIC_PATH}")
        if not example_path.exists():
            raise FileNotFoundError(f"Ejemplo no encontrado: {example_path}")

        self.rubric_text = extract_pdf_text(RUBRIC_PATH)
        self.example_text = extract_pdf_text(example_path)

        if not self.rubric_text.strip() or not self.example_text.strip():
            raise RuntimeError("No se pudo extraer texto de los PDFs de referencia.")
        
        if provider == "groq":
            self._client = OpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1")
        elif provider == "gemini":
            self._gemini_client = genai.Client(api_key=api_key)

    def correct_essay(self, student_name: str, essay_text: str) -> str:
        # Enrutador: Decide qué API usar en el momento de la llamada
        if self.provider == "gemini":
            return self._call_gemini(essay_text, student_name)
        elif self.provider == "groq":
            return self._call_groq(essay_text, student_name)
        else:
            raise ValueError(f"Proveedor desconocido: {self.provider}")

    def _call_groq(self, essay_text: str, student_name: str, retries: int = 3) -> str:
        from openai import OpenAI
        client = self._client
        user_message = self._build_prompt(essay_text, student_name)
        
        last_error = None
        for attempt in range(retries):
            try:
                response = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[
                        {"role": "system", "content": SYSTEM_INSTRUCTION},
                        {"role": "user", "content": user_message},
                    ],
                    temperature=0.1
                )
                return response.choices[0].message.content.strip()
            except Exception as e:
                last_error = e
                time.sleep(10 * (attempt + 1)) # Espera 10s, luego 20s, etc.
                
        raise RuntimeError(f"Groq API falló tras {retries} intentos. Error: {last_error}")

    def _call_gemini(self, essay_text: str, student_name: str, retries: int = 3) -> str:
        client = self._gemini_client
        full_prompt = f"{SYSTEM_INSTRUCTION}\n\n{self._build_prompt(essay_text, student_name)}"
        
        last_error = None
        for attempt in range(retries):
            try:
                response = client.models.generate_content(
                    model='gemini-2.0-flash',
                    config=genai.types.GenerateContentConfig(
                        system_instruction=SYSTEM_INSTRUCTION,
                        temperature=0.1
                    ),
                    contents=self._build_prompt(essay_text, student_name)
                )

                return response.text.strip()
            except Exception as e:
                last_error = e
                time.sleep(10 * (attempt + 1))
                
        raise RuntimeError(f"Gemini API falló tras {retries} intentos. Error: {last_error}")

    def _build_prompt(self, essay_text: str, student_name: str) -> str:
        return (
            f"Please correct the following student essay for the '{self.essay_type}' task.\n\n"
            f"=== OFFICIAL RUBRIC ===\n{self.rubric_text}\n\n"
            f"=== CORRECTION EXAMPLE ===\n{self.example_text}\n\n"
            f"Student name: {student_name}\n\n"
            f"STUDENT ESSAY:\n\"\"\"\n{essay_text}\n\"\"\""
        )


# ─────────────────────────────────────────────────────────────
# PARSING
# ─────────────────────────────────────────────────────────────

def parse_sections(full_text: str) -> dict:
    sections = {"grade": "", "errors": "", "comments": "", "corrected": ""}
    if not full_text:
        return sections
    blocks = re.split(r"(?=##\s)", full_text)
    for block in blocks:
        b = block.strip()
        if re.match(r"##\s*GRADE SUMMARY",     b, re.IGNORECASE):
            sections["grade"]     = re.sub(r"^##\s*GRADE SUMMARY\s*",     "", b, flags=re.IGNORECASE).strip()
        elif re.match(r"##\s*ERROR ANALYSIS",  b, re.IGNORECASE):
            sections["errors"]    = re.sub(r"^##\s*ERROR ANALYSIS\s*",    "", b, flags=re.IGNORECASE).strip()
        elif re.match(r"##\s*EXAMINER COMMENTS",b, re.IGNORECASE):
            sections["comments"]  = re.sub(r"^##\s*EXAMINER COMMENTS\s*", "", b, flags=re.IGNORECASE).strip()
        elif re.match(r"##\s*CORRECTED ESSAY", b, re.IGNORECASE):
            sections["corrected"] = re.sub(r"^##\s*CORRECTED ESSAY\s*",   "", b, flags=re.IGNORECASE).strip()
    return sections


def extract_total_score(text: str):
    """Devuelve int o '' si no se encuentra."""
    if not text:
        return ""
    m = re.search(r"TOTAL\s*:\s*(\d{1,2})\s*/\s*20", text, re.IGNORECASE)
    return int(m.group(1)) if m else ""


# ─────────────────────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────────────────────

STYLE_GRADE     = {"fill": PatternFill("solid", start_color="D6E4F0"), "font": Font(name="Calibri", size=11),               "align": Alignment(wrap_text=True, vertical="top")}
STYLE_NOTA      = {"fill": PatternFill("solid", start_color="FFF2CC"), "font": Font(name="Calibri", size=12, bold=True),     "align": Alignment(horizontal="center", vertical="center")}
STYLE_ERRORS    = {"fill": PatternFill("solid", start_color="FCE4D6"), "font": Font(name="Calibri", size=11),               "align": Alignment(wrap_text=True, vertical="top")}
STYLE_COMMENTS  = {"fill": PatternFill("solid", start_color="EDE7F6"), "font": Font(name="Calibri", size=11),               "align": Alignment(wrap_text=True, vertical="top")}
STYLE_CORRECTED = {"fill": PatternFill("solid", start_color="E8F5E9"), "font": Font(name="Calibri", size=11),               "align": Alignment(wrap_text=True, vertical="top")}
STYLE_FULL      = {"fill": PatternFill("solid", start_color="EBF5EB"), "font": Font(name="Calibri", size=11),               "align": Alignment(wrap_text=True, vertical="top")}

def _apply(cell, style):
    cell.fill      = style["fill"]
    cell.font      = style["font"]
    cell.alignment = style["align"]


# ─────────────────────────────────────────────────────────────
# HOJA DE RESUMEN
# ─────────────────────────────────────────────────────────────

def build_summary_sheet(wb, ws_data, essay_type: str):
    SHEET_NAME = "Resumen"
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    header_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="2E75B6")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    score_fill   = PatternFill("solid", start_color="FFF2CC")
    score_font   = Font(name="Calibri", size=11, bold=True)
    alt_fill     = PatternFill("solid", start_color="F2F7FB")

    ws.merge_cells("A1:F1")
    title           = ws["A1"]
    title.value     = f"Resumen de correcciones — {essay_type}"
    title.font      = Font(name="Calibri", size=13, bold=True, color="2E75B6")
    title.alignment = center_align
    ws.row_dimensions[1].height = 24

    headers    = ["Alumno", "TOTAL /20", "Content /5", "Comm. Ach. /5", "Organization /5", "Language /5"]
    col_widths = [28, 13, 13, 16, 16, 13]
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=2, column=ci, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 18

    pats = {
        "content":  re.compile(r"Content\s*:\s*(\d)/5",                    re.IGNORECASE),
        "comm":     re.compile(r"Communicative\s+Achievement\s*:\s*(\d)/5", re.IGNORECASE),
        "org":      re.compile(r"Organization\s*:\s*(\d)/5",               re.IGNORECASE),
        "language": re.compile(r"Language\s*:\s*(\d)/5",                   re.IGNORECASE),
    }

    row_i  = 3
    scores = []

    for row in ws_data.iter_rows(min_row=2):
        name      = str(row[COL_NOMBRE - 1].value).strip() if row[COL_NOMBRE - 1].value else ""
        full_text = str(row[COL_FULL   - 1].value).strip() if row[COL_FULL   - 1].value else ""
        if not name and not full_text:
            continue

        total = extract_total_score(full_text)
        sub   = {k: (int(p.search(full_text).group(1)) if p.search(full_text) else "") for k, p in pats.items()}

        for ci, val in enumerate([name, total, sub["content"], sub["comm"], sub["org"], sub["language"]], start=1):
            cell           = ws.cell(row=row_i, column=ci, value=val)
            cell.font      = Font(name="Calibri", size=11, bold=(ci == 2))
            cell.alignment = left_align if ci == 1 else center_align
            if ci == 2:
                cell.fill = score_fill
                cell.font = score_font
            elif row_i % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_i].height = 16

        if isinstance(total, int):
            scores.append(total)
        row_i += 1

    if scores:
        avg_row       = row_i + 1
        lbl           = ws.cell(row=avg_row, column=1, value="MEDIA DE CLASE")
        lbl.font      = Font(name="Calibri", size=11, bold=True)
        lbl.alignment = left_align
        val           = ws.cell(row=avg_row, column=2, value=round(sum(scores) / len(scores), 2))
        val.font      = Font(name="Calibri", size=12, bold=True, color="C00000")
        val.fill      = PatternFill("solid", start_color="FCE4D6")
        val.alignment = center_align

    wb.move_sheet(SHEET_NAME, offset=-len(wb.sheetnames) + 1)


# ─────────────────────────────────────────────────────────────
# LÓGICA EXCEL — opera en memoria
# ─────────────────────────────────────────────────────────────

OUTPUT_COLS = [
    (COL_GRADE,     "Grade Summary",       45, STYLE_GRADE),
    (COL_NOTA,      "Nota /20",            10, STYLE_NOTA),
    (COL_ERRORS,    "Error Analysis",      55, STYLE_ERRORS),
    (COL_COMMENTS,  "Examiner Comments",   45, STYLE_COMMENTS),
    (COL_CORRECTED, "Corrected Essay",     70, STYLE_CORRECTED),
    (COL_FULL,      "Corrección completa", 80, STYLE_FULL),
]

def _col_letter(n: int) -> str:
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def setup_headers(ws):
    # Deshacer merges en fila 1 para evitar MergedCell read-only
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= 1 <= rng.max_row:
            ws.unmerge_cells(str(rng))

    for col_idx, header, width, style in OUTPUT_COLS:
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is None:
            cell.value     = header
            cell.font      = Font(bold=True, name="Calibri", size=11)
            cell.fill      = style["fill"]
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[_col_letter(col_idx)].width = width


def write_row(ws, row_num: int, full_text: str):
    sections = parse_sections(full_text)
    score    = extract_total_score(full_text)
    for col, value, style in [
        (COL_GRADE,     sections["grade"],     STYLE_GRADE),
        (COL_NOTA,      score,                 STYLE_NOTA),
        (COL_ERRORS,    sections["errors"],    STYLE_ERRORS),
        (COL_COMMENTS,  sections["comments"],  STYLE_COMMENTS),
        (COL_CORRECTED, sections["corrected"], STYLE_CORRECTED),
        (COL_FULL,      full_text,             STYLE_FULL),
    ]:
        _apply(ws.cell(row=row_num, column=col, value=value), style)
    return score


def process_excel(
    excel_bytes: bytes,
    corrector: CambridgeCorrector,
    progress_callback=None,
    save_callback=None,
) -> tuple[io.BytesIO, dict]:
    """
    Parámetros
    ----------
    excel_bytes       : contenido del .xlsx subido por el usuario (bytes)
    corrector         : instancia de CambridgeCorrector ya inicializada
    progress_callback : función opcional f(current, total, student_name, status)
                        que app.py usa para actualizar la UI en tiempo real

    Devuelve
    --------
    (output_buffer, stats)
        output_buffer : io.BytesIO con el Excel procesado, listo para descargar
        stats         : {"success": int, "skipped": int, "failed": list[str]}
    """
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    setup_headers(ws)

    # Recoger filas a procesar
    data_rows = []
    for row in ws.iter_rows(min_row=2):
        name_cell  = row[COL_NOMBRE    - 1]
        essay_cell = row[COL_REDACCION - 1]
        full_cell  = row[COL_FULL      - 1]
        if essay_cell.value and str(essay_cell.value).strip():
            data_rows.append((essay_cell.row, name_cell, essay_cell, full_cell))

    total   = len(data_rows)
    success = 0
    skipped = 0
    failed  = []

    for idx, (row_num, name_cell, essay_cell, full_cell) in enumerate(data_rows, start=1):
        student_name = str(name_cell.value).strip() if name_cell.value else f"Alumno_fila{row_num}"
        essay_text   = str(essay_cell.value).strip()

        # ── Ya corregida: rellenar columnas desglosadas si faltan ─────────────
        if full_cell.value and str(full_cell.value).strip():
            grade_cell = ws.cell(row=row_num, column=COL_GRADE)
            if not grade_cell.value:
                write_row(ws, row_num, str(full_cell.value))
            skipped += 1
            if progress_callback:
                progress_callback(idx, total, student_name, "skipped")
            continue

        # ── Corregir ──────────────────────────────────────────────────────────
        if progress_callback:
            progress_callback(idx, total, student_name, "correcting")
        try:
            full_text = corrector.correct_essay(student_name, essay_text)
            score     = write_row(ws, row_num, full_text)
            success  += 1
            if save_callback:
                checkpoint = io.BytesIO()
                wb.save(checkpoint)
                checkpoint.seek(0)
                save_callback(checkpoint.getvalue())
            if progress_callback:
                progress_callback(idx, total, student_name, f"done:{score}")
        except Exception as e:
            failed.append(student_name)
            if progress_callback:
                progress_callback(idx, total, student_name, f"error:{e}")

        if idx < total:
            time.sleep(DELAY_BETWEEN_CALLS)

    # ── Hoja de resumen ───────────────────────────────────────────────────────
    build_summary_sheet(wb, ws, corrector.essay_type)

    # ── Serializar a memoria ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, {"success": success, "skipped": skipped, "failed": failed}
